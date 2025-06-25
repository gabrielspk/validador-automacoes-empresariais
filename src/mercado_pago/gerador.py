import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def gerar_relatorio(df):
    df['Nome Arquivo'] = df['Nome Arquivo'].str.lower()
    
    originais = df[df['Nome Arquivo'].str.contains('_0', na=False) & 
                   ~df['Status'].str.contains('Corrigido|Erro', na=False) &
                   ~df['Processo'].str.contains('MERCADO_PAGO_EXPEDICAO_BACKUP', na=False)
                   ]
    
    corrigidos = df[df['Nome Arquivo'].str.contains('_0', na=False) 
                    & df['Status'].str.contains('Corrigido', na=False)
                    ]
    
    erros = df[df['Nome Arquivo'].str.contains('_0', na=False) & df['Status'].str.contains('Erro', na=False)]

    total_originais = len(originais)
    total_corrigidos = len(corrigidos)
    total_erros = len(erros)

    concatenados = df[df['Nome Arquivo'].str.contains('conca', na=False)].sort_values(by='Nome Arquivo')
    registros_concat, registros_fpl = [], []

    for i in range(len(concatenados) - 1):
        atual, proximo = concatenados.iloc[i], concatenados.iloc[i + 1]
        if not atual['Nome Arquivo'].endswith('.fpl') and proximo['Nome Arquivo'].endswith('.fpl'):
            registros_concat.append(atual['Registros'])
            registros_fpl.append(proximo['Registros'])

    total_concatenados = sum(registros_concat)
    total_fpl = sum(registros_fpl)

    phoenix = df[df['Processo'].str.contains('PHOENIX', na=False)]
    parte2 = df[df['Processo'].str.contains('PARTE_2_CONCAT', na=False) & df['Status'].str.contains('Entregue', na=False)]
    parte3 = df[df['Processo'].str.contains('PARTE_3', na=False)].sort_values(by='Nome Arquivo')

    total_phoenix = phoenix['Registros'].sum()
    total_parte_2 = parte2['Registros'].sum()
    total_parte_3 = parte3['Registros'].groupby(parte3.index // 3).sum().sum()

    categorias = ['Total Registros Originais', 'Total Registros Concatenados', 'Total Registros FPL',
                  'Total Registros Phoenix', 'Total FPLs concatenados', 'Total FPLs enviados ao manuseio']
    quantidades = [total_originais, total_concatenados, total_fpl, total_phoenix, total_parte_2, total_parte_3]

    idx = 1
    if total_corrigidos > 0:
        categorias.insert(1, 'Registros Originais Corrigidos')
        quantidades.insert(1, total_corrigidos)
        idx += 1
    if total_erros > 0:
        categorias.insert(idx, 'Registros originais erro')
        quantidades.insert(idx, total_erros)

    return pd.DataFrame({'Categoria': categorias, 'Quantidade': quantidades})

def salvar_relatorio_com_validacao(df, caminho):
    # 1. Salva o relatório normalmente
    df.to_excel(caminho, index=False)

    # 2. Abre com openpyxl
    wb = load_workbook(caminho)
    ws = wb.active

    # 3. Cores
    vermelho = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')

    # 4. Mapear categorias para comparação
    dados = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2, max_col=2)}

    # 5. Comparações
    if 'Total Registros Concatenados' in dados and 'Total Registros FPL' in dados:
        if dados['Total Registros FPL'] < dados['Total Registros Concatenados']:
            # Pintar a célula do FPL
            for row in ws.iter_rows(min_row=2, max_col=2):
                if row[0].value == 'Total Registros FPL':
                    row[1].fill = vermelho

    if 'Total Registros Concatenados' in dados and 'Total Registros Phoenix' in dados:
        if dados['Total Registros Phoenix'] < dados['Total Registros Concatenados']:
            # Pintar Phoenix
            for row in ws.iter_rows(min_row=2, max_col=2):
                if row[0].value == 'Total Registros Phoenix':
                    row[1].fill = vermelho

    wb.save(caminho)