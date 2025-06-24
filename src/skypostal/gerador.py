import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime

def gerar_relatorio(df):
    
    data_atual = datetime.date.today()
    ano_atual = data_atual.year

    df['Nome Arquivo'] = df['Nome Arquivo'].str.lower()
    df['Processo'] = df['Processo'].str.lower()
    
    originais = df[df['Nome Arquivo'].str.contains(f'importskypostal_{ano_atual}', na=False) & ~df['Status'].str.contains('Corrigido|Erro', na=False)]
    corrigidos = df[df['Nome Arquivo'].str.contains(f'importskypostal_{ano_atual}', na=False) & df['Status'].str.contains('Corrigido', na=False)]
    erros = df[df['Nome Arquivo'].str.contains(f'importskypostal_{ano_atual}', na=False) & df['Status'].str.contains('Erro', na=False)]
    parados = df[df['Nome Arquivo'].str.contains(f'importskypostal_{ano_atual}', na=False) & df['Status'].str.contains('Parado', na=False)]
    entregar = df[df['Nome Arquivo'].str.contains(f'importskypostal_{ano_atual}', na=False) & df['Status'].str.contains('Entregar', na=False)]

    total_originais = len(originais)
    total_corrigidos = len(corrigidos)
    total_erros = len(erros)
    total_parados = len(parados)
    total_entregar = len(entregar)

    concatenados = df[df['Nome Arquivo'].str.contains('conca', na=False)].sort_values(by='Nome Arquivo')
    registros_concat, registros_fpl = [], []

    for i in range(len(concatenados) - 1):
        atual, proximo = concatenados.iloc[i], concatenados.iloc[i + 1]
        if not atual['Nome Arquivo'].endswith('.fpl') and proximo['Nome Arquivo'].endswith('.fpl'):
            registros_concat.append(atual['Registros'])
            registros_fpl.append(proximo['Registros'])

    total_concatenados = sum(registros_concat)
    total_fpl = sum(registros_fpl)

    categorias = ['Total Registros Originais', 'Total Registros Concatenados', 'Total Registros FPL']
    quantidades = [total_originais, total_concatenados, total_fpl]

    idx = 1
    if total_corrigidos > 0:
        categorias.insert(1, 'Registros Originais Corrigidos')
        quantidades.insert(1, total_corrigidos)
        idx += 1
    
    if total_erros > 0:
        categorias.insert(idx, 'Registros originais erro')
        quantidades.insert(idx, total_erros)
        idx += 1

    if total_parados > 0:
        categorias.insert(idx, 'Registros originais erro')
        quantidades.insert(idx, total_parados)
        idx += 1

    if total_entregar > 0:
        categorias.insert(idx, 'Registros originais erro')
        quantidades.insert(idx, total_entregar)

    return pd.DataFrame({'Categoria': categorias, 'Quantidade': quantidades})

def salvar_relatorio_com_validacao(df, caminho):
    # 1. Salva o relatório normalmente
    df.to_excel(caminho, index=False)

    # 2. Abre com openpyxl
    wb = load_workbook(caminho)
    ws = wb.active

    # 3. Cores
    vermelho = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')

    # 4. Mapear categorias para comparação
    dados = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2, max_col=2)}

    # 5. Comparações
    if 'Total Registros Concatenados' in dados and 'Total Registros FPL' in dados:
        if dados['Total Registros FPL'] < dados['Total Registros Concatenados']:
            # Pintar a célula do FPL
            for row in ws.iter_rows(min_row=2, max_col=2):
                if row[0].value == 'Total Registros FPL':
                    row[1].fill = vermelho

    wb.save(caminho)